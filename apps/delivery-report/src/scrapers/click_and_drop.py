"""
Click & Drop scraper — Playwright automation to:
1. Login to Royal Mail Click & Drop
2. Download the Manifested Orders XLS export
3. Parse tracking numbers matched to Amazon order references
"""

import logging
import re
import tempfile
import time
from datetime import datetime, timedelta
from pathlib import Path

from playwright.sync_api import Browser, Page, sync_playwright

from src.config import CLICK_DROP_EMAIL, CLICK_DROP_PASSWORD

log = logging.getLogger(__name__)

MANIFESTED_ORDERS_URL = "https://business.parcel.royalmail.com/reports/manifested-orders/"
LOGIN_URL_PREFIX = "auth.parcel.royalmail.com"
AMAZON_ORDER_PATTERN = re.compile(r"\d{3}-\d{7}-\d{7}")


def scrape_tracking(days: int = 28) -> dict[str, dict]:
    """
    Login to Click & Drop, download XLS export, parse Amazon tracking data.

    Returns:
        Dict of order_id -> {tracking, cd_status, despatch_date}
    """
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        try:
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            _login(page)
            xls_path = _download_xls(page)
            if xls_path:
                return _parse_xls(xls_path, days)
            else:
                log.warning("XLS download failed, falling back to DOM scraping")
                return _scrape_dom(page, days)
        finally:
            browser.close()


def _dismiss_cookie_modal(page: Page) -> None:
    """Dismiss the Tealium GDPR cookie consent modal if present."""
    try:
        # Aggressively remove the modal and ALL its overlays via JS first
        removed = page.evaluate("""() => {
            let removed = 0;
            // Remove the modal itself
            const modal = document.getElementById('__tealiumGDPRecModal');
            if (modal) { modal.remove(); removed++; }
            // Remove any privacy prompt elements that intercept clicks
            document.querySelectorAll(
                '.privacy_prompt_footer, .privacy_prompt_content, ' +
                '.privacy_prompt, [class*="privacy_prompt"], ' +
                '[class*="tealium"], [id*="tealium"], ' +
                '[class*="cookie-consent"], [class*="gdpr"]'
            ).forEach(e => { e.remove(); removed++; });
            // Remove any fixed/absolute overlays that might block clicks
            document.querySelectorAll('div[style*="position: fixed"], div[style*="position:fixed"]').forEach(e => {
                if (e.style.zIndex > 100 || e.querySelector('[class*="privacy"], [class*="cookie"], [class*="consent"]')) {
                    e.remove(); removed++;
                }
            });
            return removed;
        }""")
        if removed:
            log.info("Removed %d GDPR/cookie modal elements via JS", removed)
            time.sleep(0.5)
    except Exception as e:
        log.debug("Cookie modal dismissal: %s", e)


def _login(page: Page) -> None:
    """Navigate to manifested orders, handle login if redirected."""
    log.info("Navigating to Click & Drop manifested orders")
    page.goto(MANIFESTED_ORDERS_URL, wait_until="networkidle", timeout=30000)
    _dismiss_cookie_modal(page)

    # Check if redirected to login
    if LOGIN_URL_PREFIX not in page.url:
        log.info("Already logged in (no redirect to auth)")
        return

    log.info("Login required at: %s", page.url)
    _dismiss_cookie_modal(page)
    time.sleep(1)

    # Log visible form elements for debugging
    inputs = page.locator("input").all()
    log.info("Found %d input fields: %s", len(inputs), [
        f"{i.get_attribute('type') or 'text'}[{i.get_attribute('name') or i.get_attribute('id') or '?'}]"
        for i in inputs[:6]
    ])

    # Fill email — try multiple selectors
    email_filled = False
    for sel in ['input[type="email"]', 'input[name="email"]', '#email', 'input[name="username"]', '#username']:
        el = page.locator(sel)
        if el.count() > 0 and el.first.is_visible():
            el.first.fill(CLICK_DROP_EMAIL)
            email_filled = True
            log.info("Filled email via: %s", sel)
            break

    if not email_filled:
        log.error("Could not find email input field")

    # Fill password — try multiple selectors
    pw_filled = False
    for sel in ['input[type="password"]', 'input[name="password"]', '#password']:
        el = page.locator(sel)
        if el.count() > 0 and el.first.is_visible():
            el.first.fill(CLICK_DROP_PASSWORD)
            pw_filled = True
            log.info("Filled password via: %s", sel)
            break

    if not pw_filled:
        # Might be a multi-step login — click Next/Continue first
        log.info("No password field visible, checking for Next/Continue button")
        for sel in ['button:has-text("Next")', 'button:has-text("Continue")', 'input[type="submit"]']:
            btn = page.locator(sel)
            if btn.count() > 0 and btn.first.is_visible():
                btn.first.click()
                log.info("Clicked '%s', waiting for password step", sel)
                time.sleep(3)
                _dismiss_cookie_modal(page)
                break

        # Retry password
        for sel in ['input[type="password"]', 'input[name="password"]', '#password']:
            el = page.locator(sel)
            if el.count() > 0 and el.first.is_visible():
                el.first.fill(CLICK_DROP_PASSWORD)
                pw_filled = True
                log.info("Filled password via: %s (after next step)", sel)
                break

    # Dismiss cookie modal again right before clicking sign-in
    _dismiss_cookie_modal(page)

    # Click sign in (force=True to bypass any remaining overlays)
    clicked = False
    for sel in ['button:has-text("Sign In")', 'button:has-text("Sign in")', 'button:has-text("Log in")', 'input[type="submit"]', 'button[type="submit"]']:
        btn = page.locator(sel)
        if btn.count() > 0 and btn.first.is_visible():
            btn.first.click(force=True)
            clicked = True
            log.info("Clicked sign-in via: %s", sel)
            break

    if not clicked:
        log.error("Could not find sign-in button")

    # Wait for redirect back to manifested orders (longer timeout)
    try:
        page.wait_for_url(f"**{MANIFESTED_ORDERS_URL}**", timeout=30000)
        log.info("Login successful, waiting for page content to load")
        # SPA keeps making requests so networkidle never fires; use domcontentloaded
        # plus explicit sleep for JS rendering
        page.wait_for_load_state("domcontentloaded", timeout=15000)
        time.sleep(5)  # Allow SPA to fetch and render manifested orders data
        log.info("Page loaded. Current URL: %s", page.url)
    except Exception as e:
        if "Timeout" in str(e) and MANIFESTED_ORDERS_URL.rstrip("/") in page.url:
            # URL redirect succeeded but load state timed out — continue anyway
            log.warning("Load state timed out but URL looks correct, continuing")
            time.sleep(3)
        else:
            log.error("Login redirect failed. Current URL: %s", page.url)
            body_text = page.inner_text("body")[:500]
            log.error("Page text: %s", body_text)
            raise


def _download_xls(page: Page) -> Path | None:
    """Click the Export/Download button, handle the date-range modal, and save the XLS.

    The Click & Drop "Export to XLS" button opens a modal dialog with a date-range
    selector (defaulting to "Last 30 days") and a submit button. We must:
    1. Click the initial "Export to XLS" button on the page
    2. Wait for the modal to appear in #modal-root
    3. Click the "Export to XLS" submit button inside the modal
    4. Capture the resulting download
    """
    log.info("Looking for XLS export button")

    # Dismiss GDPR modal that reappears on the manifested orders page
    _dismiss_cookie_modal(page)

    # Wait for the table or export controls to appear (SPA may still be loading)
    try:
        page.wait_for_selector("table, button:has-text('Export')", timeout=15000)
        time.sleep(1)
    except Exception:
        log.warning("Timed out waiting for table/export controls")

    _dismiss_cookie_modal(page)

    # Step 1: Click the "Export to XLS" button on the main page
    btn = page.locator('button:has-text("Export to XLS")')
    if btn.count() == 0:
        log.warning("No 'Export to XLS' button found on page")
        body_text = page.inner_text("body")[:800]
        log.warning("Page text preview: %s", body_text)
        return None

    log.info("Clicking 'Export to XLS' button to open modal")
    btn.first.click()

    # Step 2: Wait for the export modal to appear in #modal-root
    try:
        page.wait_for_selector("#modal-root #modal-panel", timeout=10000)
        log.info("Export modal appeared")
    except Exception:
        log.warning("Export modal did not appear after clicking button")
        return None

    time.sleep(1)  # Allow modal to fully render

    # Log modal content for debugging
    modal_text = page.evaluate("""() => {
        const root = document.getElementById('modal-root');
        return root ? root.textContent.trim().substring(0, 300) : '';
    }""")
    log.info("Modal text: %s", modal_text)

    # Step 3: Click the submit "Export to XLS" button inside the modal
    modal_submit = page.locator('#modal-root button[type="submit"]')
    if modal_submit.count() == 0:
        # Fallback: any button with "Export" text inside the modal
        modal_submit = page.locator('#modal-root button:has-text("Export")')

    if modal_submit.count() == 0:
        log.warning("No submit button found in export modal")
        return None

    log.info("Clicking modal submit button to trigger download")
    try:
        with page.expect_download(timeout=60000) as download_info:
            modal_submit.first.click()
        download = download_info.value
        tmp = Path(tempfile.mkdtemp()) / download.suggested_filename
        download.save_as(str(tmp))
        log.info("Downloaded XLS: %s (%d bytes)", tmp.name, tmp.stat().st_size)
        return tmp
    except Exception as e:
        log.error("Download failed after clicking modal submit: %s", e)
        return None


def _read_xls_xlrd(xls_path: Path) -> list[tuple]:
    """Read old-format .xls file using xlrd. Returns list of row tuples."""
    import xlrd

    wb = xlrd.open_workbook(str(xls_path))
    ws = wb.sheet_by_index(0)
    rows = []
    for i in range(ws.nrows):
        row_values = []
        for j in range(ws.ncols):
            cell = ws.cell(i, j)
            # Convert xlrd date cells to datetime objects
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt_tuple = xlrd.xldate_as_tuple(cell.value, wb.datemode)
                    row_values.append(datetime(*dt_tuple))
                except Exception:
                    row_values.append(cell.value)
            else:
                row_values.append(cell.value)
        rows.append(tuple(row_values))
    return rows


def _read_xlsx_openpyxl(xls_path: Path) -> list[tuple]:
    """Read .xlsx file using openpyxl. Returns list of row tuples."""
    import openpyxl

    wb = openpyxl.load_workbook(str(xls_path), read_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _parse_xls(xls_path: Path, days: int = 28) -> dict[str, dict]:
    """Parse the Click & Drop XLS export.

    Uses xlrd for .xls (old binary format) and openpyxl for .xlsx.
    Click & Drop exports old-format .xls files.
    """
    log.info("Parsing XLS: %s", xls_path)

    suffix = xls_path.suffix.lower()
    if suffix == ".xls":
        rows = _read_xls_xlrd(xls_path)
    else:
        rows = _read_xlsx_openpyxl(xls_path)

    if not rows:
        log.warning("XLS file is empty")
        return {}

    headers = [str(h or "").strip() for h in rows[0]]

    # Find column indexes
    def find_col(names: list[str]) -> int | None:
        for name in names:
            name_lower = name.lower()
            for i, h in enumerate(headers):
                if h.lower() == name_lower:
                    return i
        return None

    channel_idx = find_col(["Channel"])
    ref_idx = find_col(["Channel reference"])
    despatch_idx = find_col(["Despatch date"])
    tracking_idx = find_col(["Tracking number"])
    status_idx = find_col(["Tracking status"])

    if ref_idx is None or tracking_idx is None:
        log.error("Required columns not found in XLS. Headers: %s", headers)
        return {}

    cutoff = datetime.now() - timedelta(days=days)
    results = {}

    for row in rows[1:]:
        # Filter for Amazon orders
        if channel_idx is not None:
            channel = str(row[channel_idx] or "")
            if "Amazon" not in channel:
                continue

        # Parse order reference
        ref = str(row[ref_idx] or "").strip()
        if not AMAZON_ORDER_PATTERN.match(ref) or ref == "*":
            continue

        # Parse despatch date
        if despatch_idx is not None:
            raw_date = row[despatch_idx]
            if isinstance(raw_date, datetime):
                despatch = raw_date.date()
            elif isinstance(raw_date, str):
                try:
                    despatch = datetime.strptime(raw_date[:10], "%Y-%m-%d").date()
                except ValueError:
                    continue
            else:
                continue

            if despatch < cutoff.date():
                continue

            despatch_str = despatch.strftime("%Y-%m-%d")
        else:
            despatch_str = None

        tracking = str(row[tracking_idx] or "").strip()
        cd_status = str(row[status_idx] or "").strip() if status_idx is not None else ""

        if tracking:
            results[ref] = {
                "tracking": tracking,
                "cd_status": cd_status,
                "despatch_date": despatch_str,
            }

    log.info("Parsed %d Amazon orders with tracking from XLS", len(results))
    return results


def _scrape_dom(page: Page, days: int = 28) -> dict[str, dict]:
    """Fallback: scrape the HTML table (page 1 only, up to 500 rows)."""
    log.info("Falling back to DOM scraping")

    # Wait for table with actual rows to load (SPA may still be rendering)
    try:
        page.wait_for_selector("table tbody tr", timeout=20000)
    except Exception:
        log.warning("No table rows found after 20s wait")
        # Log what IS on the page
        tables = page.locator("table").all()
        log.info("Found %d <table> elements", len(tables))
        body_text = page.inner_text("body")[:1000]
        log.info("Page text: %s", body_text)
    time.sleep(2)

    rows = page.locator("table tbody tr").all()
    log.info("Found %d table rows", len(rows))
    results = {}

    for row in rows:
        cells = row.locator("td").all()
        if len(cells) < 6:
            continue

        texts = [c.inner_text().strip() for c in cells]

        # Column 0 = Channel, Column 1 = Order ref, Column ~5 = Tracking
        channel = texts[0] if len(texts) > 0 else ""
        if "Amazon" not in channel:
            continue

        ref_match = AMAZON_ORDER_PATTERN.search(texts[1] if len(texts) > 1 else "")
        if not ref_match:
            continue
        ref = ref_match.group(0)

        tracking = texts[-2] if len(texts) >= 2 else ""  # Tracking is usually second-to-last
        if tracking and len(tracking) > 5:
            results[ref] = {
                "tracking": tracking,
                "cd_status": texts[-1] if len(texts) >= 1 else "",
                "despatch_date": None,
            }

    log.info("Scraped %d Amazon orders from DOM", len(results))
    return results
